
from configparser import ConfigParser
import eel
import json
import os
import win32com.client as win32
from openpyxl import load_workbook
from collections import OrderedDict

template =None
template_file =None
current =None
json_list = []
cube_list = []

## 설정파일 읽기
parser = ConfigParser()
parser.read('config.ini')

############################################################
## LOAD JSON FILE INTER FACE
############################################################
@eel.expose
def load_json_file(json_file_name):
    global template_file
    with open(parser.get('settings', 'data')+"/"+json_file_name, "rb") as fin:
        template_file = json.load(fin)
        eel.setup_template_file(template_file)

@eel.expose
def load_data_file(file_name):
    global template_file,load_wb
    #data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
    print('--------- load data file ---------'+os.getcwd())
    fname = parser.get('settings', 'data')+"/"+file_name

    # 파일확장자 체크 xls 일경우 xlsx cervert
    if file_name.endswith(".xls"):
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(os.getcwd() + fname)
        wb.SaveAs(os.getcwd() + fname + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
        wb.Close()
        excel.Application.Quit()
        load_wb = load_workbook(fname + "x", data_only=True)
        os.remove(os.getcwd() + fname + "x") # 파일변경작업 후 파일삭제
    else:
        load_wb = load_workbook(fname, data_only=True)

    ws = load_wb.worksheets
    final = []
    for sheet in ws:
        #sheet1 = load_wb.get_sheet_by_name('Sheet1')
        # get_active_sheet()로 활성화된 시트를 불러올 수도 있습니다.
        # sheet2 = excelFile.get_active_sheet()
        res = list(sheet)
        obj ={}
        rows = []
        for x in range(1, len(res)):
            partFinal = OrderedDict()
            for i in range(0,14):
                if res[0][i].value == 'id':
                    obj["id"]= res[x][i].value
                elif res[0][i].value == 'name':
                    obj["name"]= res[x][i].value
                elif res[0][i].value == 'main_image':
                    obj["main_image"]= res[x][i].value
                elif res[0][i].value in ['pos','pos_movie']:
                    _arr = res[x][i].value.split("|")
                    for k in range(0, len(_arr)):
                        _arr[k] = _arr[k].split(",")
                        _arr[k] = list(map(int, _arr[k]))
                    partFinal[res[0][i].value] = _arr
                else:    
                    partFinal[res[0][i].value] = res[x][i].value
            # TODO : setting array
            rows.append(partFinal)
        obj["template"] = rows
        final.append(obj)

    j = json.dumps(final) 
    template_file = json.loads(j)
    eel.setup_template_file(template_file)

