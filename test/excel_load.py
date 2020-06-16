from openpyxl import load_workbook
import json
from collections import OrderedDict

 
#data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook("data/용산돌산치.xlsx", data_only=True)
#시트 이름으로 불러오기
#load_ws = load_wb['Sheet1']
sheet1 = load_wb.get_sheet_by_name('Sheet1')
# get_active_sheet()로 활성화된 시트를 불러올 수도 있습니다.
# sheet2 = excelFile.get_active_sheet()
res = list(sheet1)
final = []
for x in range(1, len(res)):
    partFinal = OrderedDict()
    partFinal[res[0][0].value] = res[x][0].value
    partFinal[res[0][1].value] = res[x][1].value
    partFinal[res[0][2].value] = res[x][2].value
    final.append(partFinal)

j = json.dumps(final) 
print(type(j))

jj = json.loads(j)
print(jj)

#for x in final:
#    print (x)


"""
allList = []
#for row in sheet1.iter_rows(min_row=2, max_row=2, min_col=1, max_col=5):
for row in sheet1.iter_rows(min_row=2):
    a = []
    print(row[1])
    for cell in row:
        #print(cell.value)
        a.append(cell.value)
    allList.append(a)
"""

"""
sh = wb.sheet_by_index(0)
# List to hold dictionaries
data_list = []
# Iterate through each row in worksheet and fetch values into dict
for rownum in range(1, sh.nrows):
    data = OrderedDict()
    row_values = sh.row_values(rownum)
    data['pattern'] = row_values[0]
    data['response'] = row_values[1]
    data_list.append(data)
# Serialize the list of dicts to JSON
j = json.dumps(data_list)
# Write to file
with open('data1.json', 'w') as f:
    f.write(j)

"""

"""

#셀 주소로 값 출력
print(load_ws['A1'].value)
 
#셀 좌표로 값 출력
print(load_ws.cell(1,2).value)


print('\n-----지정한 셀 출력-----')
get_cells = load_ws['A1':'D2']
for row in get_cells:
        for cell in row:
            print(cell.value)

print('\n-----모든 행 단위로 출력-----')
for row in load_ws.rows:
    print(row)

print('\n-----모든 열 단위로 출력-----')
for column in load_ws.columns:
    print(column)


print('\n-----모든 행과 열 출력-----')
all_values = []
for row in load_ws.rows:
    row_value = []
    for cell in row:
        row_value.append(cell.value)
    all_values.append(row_value)
print(all_values)

"""

